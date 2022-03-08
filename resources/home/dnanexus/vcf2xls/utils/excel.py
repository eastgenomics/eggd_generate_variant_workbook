from pathlib import Path
from string import ascii_uppercase as uppercase
from unicodedata import name

import Levenshtein as levenshtein
from openpyxl.styles import Border, DEFAULT_FONT, Font, Side
from openpyxl.styles.fills import PatternFill
import pandas as pd

# openpyxl style settings
THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DEFAULT_FONT.name = 'Calibri'

class excel():
    """
    Functions for wrangling variant data into spreadsheet formatting and
    writing output file

    Attributes
    ----------
    args : argparse.Namespace
        arguments passed from command line
    vcfs : list of pd.DataFrame
        list of dataframes formatted to write to file from vcf() methods
    refs : list
        list of reference names parsed from vcf headers
    writer : pandas.io.excel._openpyxl.OpenpyxlWriter
        writer object for writing Excel data to file
    workbook : openpyxl.workbook.workbook.Workbook
        openpyxl workbook object for interacting with per-sheet writing and
        formatting of output Excel file

    Outputs
    -------
    {args.output}.xlsx : file
        Excel file with variants written to, name passed from command line or
        inferred from input vcf name if not specified
    """
    def __init__(self, args, vcfs, refs) -> None:
        print(f"Writing to output file: {Path(args.output).absolute()}")
        self.args = args
        self.vcfs = vcfs
        self.refs = refs
        self.writer = pd.ExcelWriter(args.output, engine='openpyxl')
        self.workbook = self.writer.book


    def generate(self) -> None:
        """
        Calls all methods in excel() to generate output file
        """
        self.write_summary()
        self.write_variants()

        self.workbook.save(self.args.output)
        print('Done!')


    def write_summary(self) -> None:
        """
        Write summary sheet to excel file
        """
        print('Writing summary sheet')
        if self.args.summary == 'dias':
            # generate summary sheet in format for RD/dias
            self.summary = self.workbook.create_sheet('summary')
            self.dias_summary()


    def dias_summary(self) -> None:
        """
        Write summary sheet in format for RD group, adds the following info:
            - sample ID, panel(s), run IDs etc.
            - formatted tables for them to fill in reporting
        """
        # write titles for summary values
        self.summary.cell(1, 1).value = "Sample ID:"
        self.summary.cell(1, 5).value = "Clinical Indication(s):"
        self.summary.cell(2, 5).value = "Panel(s):"
        self.summary.cell(40, 1).value = "Reads:"
        self.summary.cell(41, 1).value = "Usable Reads:"
        self.summary.cell(43, 1).value = "Workflow:"
        self.summary.cell(44, 1).value = "Workflow ID:"

        # write summary values
        self.summary.cell(1, 2).value = self.args.sample
        self.summary.cell(1, 6).value = self.args.clinical_indication
        self.summary.cell(2, 6).value = self.args.panel
        self.summary.cell(40, 2).value = self.args.reads
        self.summary.cell(41, 2).value = self.args.usable_reads
        self.summary.cell(43, 2).value = self.args.workflow[0]
        self.summary.cell(44, 2).value = self.args.workflow[1]

        # write args passed to script to generate report
        self.summary.cell(46, 1).value = "Filters applied:"
        if self.args.filter:
            for idx, filter in enumerate(self.args.filter):
                self.summary.cell(46 + idx, 2).value = filter
        else:
            self.summary.cell(46, 2).value = "None"

        # write center reporting section tables
        self.summary.cell(9, 2).value = "Phenotype:"

        self.summary.cell(16, 2).value = "Panels"
        self.summary.cell(16, 3).value = "Excel file"
        self.summary.cell(16, 4).value = "Comments"
        self.summary.cell(16, 6).value = "Analysis by"
        self.summary.cell(16, 7).value = "Date"
        self.summary.cell(16, 8).value = "Checked by"
        self.summary.cell(16, 9).value = "Date"

        self.summary.cell(21, 2).value = "Sanger sequencing confirmation"
        self.summary.cell(22, 2).value = "Gene"
        self.summary.cell(22, 3).value = "NM_#"
        self.summary.cell(22, 4).value = "Coordinate"
        self.summary.cell(22, 5).value = "cDNA"
        self.summary.cell(22, 6).value = "Protein change"
        self.summary.cell(22, 7).value = "WS#"
        self.summary.cell(22, 8).value = "Confirmed (Y/N)"

        self.summary.cell(28, 2).value = "GEM comments summary"
        self.summary.cell(28, 4).value = "Date"

        # merge some title columns that have longer text
        self.summary.merge_cells(
            start_row=9, end_row=9, start_column=2, end_column=5)
        self.summary.merge_cells(
            start_row=21, end_row=21, start_column=2, end_column=8)
        self.summary.merge_cells(
            start_row=16, end_row=16, start_column=4, end_column=5)
        self.summary.merge_cells(
            start_row=28, end_row=28, start_column=2, end_column=3)
        self.summary.merge_cells(
            start_row=28, end_row=28, start_column=4, end_column=6)

        # set titles to bold
        title_cells = [
            "A1", "A40", "A41", "A43", "A44", "B1", "B9", "B16",
            "B21", "B22", "B28", "B40", "B41", "B43", "B44", "C16",
            "C22", "D16", "D22", "D28", "E1", "E2", "E22", "F1", "F2",
            "F16", "F22", "G16", "G22", "H16", "H22", "I16"
        ]
        for cell in title_cells:
            self.summary[cell].font = Font(bold=True, name=DEFAULT_FONT.name)

        # set column widths for readability
        self.summary.column_dimensions['A'].width = 13
        self.summary.column_dimensions['B'].width = 13
        self.summary.column_dimensions['C'].width = 13
        self.summary.column_dimensions['D'].width = 13
        self.summary.column_dimensions['E'].width = 18
        self.summary.column_dimensions['F'].width = 16
        self.summary.column_dimensions['G'].width = 16
        self.summary.column_dimensions['H'].width = 16

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="0CABA8")

        colour_cells = [
            "B9", "B16", "B21", "B22", "B28", "C16", "C22", "D16", "D22",
            "D28", "E22", "F16", "F22", "G16", "G22", "H16", "H22", "I16"
        ]
        for cell in colour_cells:
            self.summary[cell].fill = blueFill

        # set borders around table areas
        row_ranges = [
            'B9:E9', 'B10:E10', 'B11:E11', 'B12:E12', 'B13:E13',
            'B16:I16', 'B17:I17', 'B18:I18',
            'B21:H21', 'B22:H22', 'B23:H23', 'B24:H24', 'B25:H25',
            'B28:F28', 'B29:F29', 'B30:F30', 'B31:F31', 'B32:F32'
        ]
        for row in row_ranges:
            for cells in self.summary[row]:
                for cell in cells:
                    cell.border = THIN_BORDER


    def write_variants(self) -> None:
        """
        Writes all variants from dataframe(s) to sheet(s) specified in
        self.args.sheets.

        If sheet names not specified, these will be set as "variant" where one
        dataframe is being written, or the vcf filename prefix if there are
        more than one dataframes to write
        """
        total_rows = sum([len(x) for x in self.vcfs])
        print(f"Writing {total_rows} rows to output xlsx file")
        with self.writer:
            # add variants
            for sheet, vcf in zip(self.args.sheets, self.vcfs):
                vcf.to_excel(
                    self.writer, sheet_name=sheet, index=False
                )
                curr_worksheet = self.writer.sheets[sheet]
                self.set_widths(curr_worksheet, vcf.columns)
                self.set_font(curr_worksheet)
                self.colour_hyperlinks(curr_worksheet)

                self.workbook.save(self.args.output)


    def set_font(self, worksheet) -> None:
        """
        Set font to all cells in variant sheet to Calibri

        Default is Times New Roman and it is ugly

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        """
        for cells in worksheet.rows:
            for cell in cells:
                cell.font = Font(name=DEFAULT_FONT.name)


    def colour_hyperlinks(self, worksheet) -> None:
        """
        Set text colour to blue if text contains hyperlink

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        """
        for ws in self.workbook:
            for cells in ws.rows:
                for cell in cells:
                    if 'HYPERLINK' in str(cell.value):
                        cell.font = Font(color='00007f', name=DEFAULT_FONT.name)


    def set_widths(self, current_sheet, sheet_columns) -> None:
        """
        Set widths for variant sheets off common names to be more readable,
        calls get_closest_match() method to determine appropriate column
        width to use.

        Parameters
        ----------
        current_sheet : openpyxl.Writer
            writer object for current sheet
        sheet_columns : list
            column names for sheet from DataFrame.columns
        """
        widths = {
            "chrom": 8,
            "pos": 12,
            "ref": 10,
            "alt": 10,
            "qual": 10,
            "af": 6,
            "dp": 10,
            'ac': 10,
            'af': 10,
            'an': 10,
            'dp': 10,
            'baseqranksum': 15,
            'clippingranksum': 16,
            "symbol": 12,
            "exon": 9,
            "variant class": 15,
            "consequence": 25,
            "hgvsc": 27,
            "hgvsp": 27,
            "gnomad": 13,
            "existing variation": 18,
            "clinvar": 10,
            "clinvar clndn": 18,
            "clinvar clinsig": 18,
            "cosmic": 15,
            "feature": 17
        }

        # generate list of 260 potential xlsx columns from A,B,C...IX,IY,IZ
        # allows for handling a lot of columns
        column_list = [
            f"{x}{y}" for x in [
                '', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'
            ] for y in uppercase
        ]

        for idx, column in enumerate(sheet_columns):
            # loop over column names, select width by closest match from dict
            # and set width in sheet letter
            width = self.get_closest_match(column.lower(), widths)
            current_sheet.column_dimensions[column_list[idx]].width = width


    def get_closest_match(self, column, widths) -> int:
        """
        Given a column name, find the closest match (if there is one) in the
        widths dict and return its width value to set. Using imprecise name
        matching as columns can differ between variant callers, annotation in
        VEP and renaming in vcf.rename_columns()

        Parameters
        ----------
        column : str
            name of column
        widths : dict
            dict of common column names and widths

        Returns
        -------
        width : int
            column width value to set
        """
        distances = {x: levenshtein.distance(column, x) for x in widths.keys()}
        closest_match = min(distances, key=distances.get)

        if distances[closest_match] <= 5:
            # close enough match to probably be correct
            width = widths[closest_match]
        else:
            # no close matches to name, use default width
            width = 13

        return width

from collections import defaultdict
import json
import operator
import os
from pathlib import Path
import re
from string import ascii_uppercase as uppercase
from timeit import default_timer as timer
from typing import Union

from colour import Color
import Levenshtein as levenshtein
import numpy as np
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl import drawing, load_workbook
from openpyxl.styles import Alignment, Border, DEFAULT_FONT, Font, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection
import pandas as pd

from .utils import is_numeric

# openpyxl style settings
THIN = Side(border_style="thin", color="000000")
MEDIUM = Side(border_style="medium", color="000001")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DEFAULT_FONT.name = 'Calibri'

# row and col counts that are to be unlocked next to
# populated table in all sheets if it is dias pipeline
# required for 'lock_sheet' function
ROW_TO_UNLOCK = 500
COL_TO_UNLOCK = 200


class excel():
    """
    Functions for wrangling variant data into spreadsheet formatting and
    writing output file

    Attributes
    ----------
    args : argparse.Namespace
        arguments passed from command line
    vcfs : list of pd.DataFrame
        list of dataframes formatted to write to file from vcf() methods
    additional_files : dict
        (optional) if addition files have been passed, dict will be populated
        with worksheet name : df of file data
    refs : list
        list of reference names parsed from vcf headers
    writer : pandas.io.excel._openpyxl.OpenpyxlWriter
        writer object for writing Excel data to file
    workbook : openpyxl.workbook.workbook.Workbook
        openpyxl workbook object for interacting with per-sheet writing and
        formatting of output Excel file

    Outputs
    -------
    {args.output}.xlsx : file
        Excel file with variants written to, name passed from command line or
        inferred from input vcf name if not specified
    """
    def __init__(self, args, vcfs, additional_files, refs) -> None:
        print(f"Writing to output file: {Path(args.output).absolute()}")
        self.args = args
        self.vcfs = vcfs
        self.additional_files = additional_files
        self.refs = refs
        self.writer = pd.ExcelWriter(args.output, engine='openpyxl')
        self.workbook = self.writer.book
        self.summary = None


    def generate(self) -> None:
        """
        Calls all methods in excel() to generate output file
        """
        self.write_summary()
        if self.args.acmg:
            for i in range(1, self.args.acmg+1):
                self.write_reporting_template(i)
        self.write_variants()
        self.write_additional_files()
        self.write_images()

        self.workbook.save(self.args.output)
        if self.args.acmg and self.args.lock_sheet:
            self.protect_rename_sheets()
        if self.args.acmg:
            self.drop_down()
        print('Done!')


    def write_summary(self) -> None:
        """
        Write summary sheet to excel file
        """
        print('Writing summary sheet')
        if self.args.summary:
            self.summary = self.workbook.create_sheet('summary')

        if self.args.summary == 'helios':
            # add summary sheet for TSO500/Helios
            self.helios_summary()
        if self.args.summary == 'dias':
            # generate summary sheet in format for RD/Dias
            self.dias_summary()


    def summary_sheet_cell_colour_key(self, row_count, to_bold) -> Union[int, list]:
        """
        Write conditions and colours of colouring applied to cells to
        the summary sheet if --colour specified

        Parameters
        ----------
        row_count : int
            counter of current row to write to in sheet
        to_bold : list
            list of cells to set to bold

        Returns
        -------
        int
            counter of current row to write to in sheet
        list
            list of cells to set to bold
        """
        # build a dict of each column and all its colour conditions
        cols_to_colours = defaultdict(dict)
        for i in self.args.colour:
            column, condition, colour = i.split(':')
            cols_to_colours[column][condition] = colour

        self.summary.cell(row_count, 1).value = "Cell colouring applied:"
        to_bold.append(f"A{row_count}")
        self.summary[f"A{row_count}"].font = Font(
            bold=True, name=DEFAULT_FONT.name
        )

        colour_col = 2
        max_colour_rows_written = 0

        # write colouring applied to each field as separate column in summary
        for column, conditions in cols_to_colours.items():
            colour_row = row_count + 1
            column_letter = get_column_letter(colour_col)

            self.summary.cell(row_count, colour_col).value = column
            to_bold.append(f"{column_letter}{row_count}")

            for condition, colour in conditions.items():
                condition = condition.replace('&', ' & ').replace('|', ' | ')
                self.summary.cell(colour_row, colour_col).value = condition
                self.summary.cell(colour_row, colour_col).data_type = 's'

                colour = self.convert_colour(colour)

                self.summary[f"{column_letter}{colour_row}"].fill = PatternFill(
                    patternType="solid",
                    start_color=colour
                )
                colour_row += 1

            # set width to wider than max value in cell
            width = max([len(x) for x in conditions.keys()])
            width = 10 if width < 13 else width
            self.summary.column_dimensions[column_letter].width = width + 3

            colour_col += 2

            if colour_row > max_colour_rows_written:
                max_colour_rows_written = colour_row

        return max_colour_rows_written, to_bold


    def helios_summary(self) -> None:
        """
        Writes summary sheet for helios pipeline with metrics such as
        variant records per sheet, dx file IDs and parameters specified
        """
        # track what cells to make bold
        to_bold = []

        # write titles for summary values
        self.summary.cell(1, 1).value = "Sample ID:"
        self.summary.cell(4, 1).value = "Name"
        self.summary.cell(5, 1).value = "Clinical indication"
        self.summary.cell(6, 1).value = "Additional comments"

        self.summary.cell(9, 1).value = "Variant totals"

        to_bold.extend(["A1", "A2", "A4", "A5", "A6", "A9"])

        # get sample name from vcf, should only be one but handle everything
        # list-wise just in case
        sample = [
            Path(x).name.replace('.vcf', '').replace('.gz', '')
            for x in self.args.vcfs
        ]
        sample = [x.split('_')[0] if '_' in x else x for x in sample]
        sample = str(sample).strip('[]').strip("'")
        self.summary.cell(1, 2).value = sample

        # split sample name into constituent parts on '-' and write to
        # separate cells for ease of them copying
        for idx, part in enumerate(sample.split('-')):
            self.summary.cell(2, idx+2).value = part

        self.summary.column_dimensions['A'].width = 36
        self.summary.column_dimensions['B'].width = 16
        self.summary.column_dimensions['C'].width = 16
        self.summary.column_dimensions['D'].width = 16

        self.summary.merge_cells(
            start_row=1, end_row=1, start_column=2, end_column=6)

        row_count = 9

        # write counts of variants
        for sheet, vcf in zip(self.args.sheets, self.vcfs):
            self.summary.cell(row_count, 2).value = sheet
            self.summary.cell(row_count, 3).value = len(vcf.index)
            to_bold.append(f"B{row_count}")
            row_count += 1

        row_count += 3

        # Parsing of MetricsOutput metrics into summary sheet
        for _, df in self.additional_files.items():
            if df.empty:
                continue

            if df.iloc[0].iloc[0] == 'Metric (UOM)':
                # its a metrics output file
                if not len(df.columns.tolist()) == 4:
                    # not 4 cols => didn't parse out just sample values in
                    # utils.parse_metrics => skip
                    continue

                # specific metrics lines we want to parse out
                idxs = []
                idxs.append(df[0].eq('Metric (UOM)').idxmax())
                idxs.append(df[0].eq('CONTAMINATION_SCORE (NA)').idxmax())
                idxs.append(df[0].eq('CONTAMINATION_P_VALUE (NA)').idxmax())
                idxs.append(df[0].eq('PCT_EXON_50X (%)').idxmax())
                idxs.append(df[0].eq('PCT_EXON_100X (%)').idxmax())

                colouring = {"green": [], "amber": [], "red": []}

                for file_row, idx in enumerate(idxs):
                    title = df.iloc[idx].iloc[0]
                    lsl = df.iloc[idx].iloc[1]
                    usl = df.iloc[idx].iloc[2]
                    sample = df.iloc[idx].iloc[3]

                    self.summary.cell(row_count, 1).value = title
                    self.summary.cell(row_count, 2).value = lsl
                    self.summary.cell(row_count, 3).value = usl
                    self.summary.cell(row_count, 4).value = sample

                    # perform colouring like in self.colour_metrics(), lazily
                    # catch anything in case of weird values to not break
                    try:
                        if title == 'CONTAMINATION_SCORE (NA)':
                            if float(sample) > float(usl):
                                colouring["amber"].append(row_count)
                            else:
                                colouring["green"].append(row_count)
                        elif title == 'CONTAMINATION_P_VALUE (NA)':
                            if float(sample) > float(usl):
                                colouring["red"].append(row_count)
                            else:
                                colouring["green"].append(row_count)
                        elif title == 'PCT_EXON_50X (%)':
                            if float(sample) >= 95:
                                colouring["green"].append(row_count)
                            else:
                                colouring["red"].append(row_count)
                        elif title == 'PCT_EXON_100X (%)':
                            if float(sample) >= 90:
                                colouring["green"].append(row_count)
                            else:
                                colouring["red"].append(row_count)
                    except Exception as err:
                        print(
                            "WARNING: error in colouring metrics values in "
                            f"summary sheet: {err}.\nContinuing without colouring"
                        )

                    to_bold.append(f"A{row_count}")
                    row_count += 1

                # do the colouring
                for colour, idxs in colouring.items():
                    for idx in idxs:
                        if colour == 'green':
                            self.summary[f"D{idx}"].fill = PatternFill(
                                patternType="solid",
                                start_color='008100'
                            )
                        elif colour == 'amber':
                             self.summary[f"D{idx}"].fill = PatternFill(
                                patternType="solid",
                                start_color='ff9f00'
                            )
                        elif colour == 'red':
                            self.summary[f"D{idx}"].fill = PatternFill(
                                patternType="solid",
                                start_color='b30000'
                            )
        row_count += 2

        # Parsing of TMB/MSI/Gene Amplifications into summary
        for _, df in self.additional_files.items():
            if df.empty:
                continue

            if df.iloc[0].iloc[0] == '[TMB]':
                for _, row in df.iterrows():
                    self.summary.cell(row_count, 1).value = row[0]
                    self.summary.cell(row_count, 2).value = row[1]
                    if row[0] != 'NA':
                        to_bold.append(f"A{row_count}")

                    row_count += 1

        row_count += 1

        # write genome reference(s) parsed from vcf header
        if self.refs:
            self.summary.cell(row_count, 1).value = "Reference:"
            self.summary[f"A{row_count}"].font = Font(
                bold=True, name=DEFAULT_FONT.name
            )
            for ref in list(set(self.refs)):
                self.summary.cell(row_count, 2).value = ref
                row_count += 1

            row_count += 2

        if self.args.human_filter:
            self.summary.cell(row_count, 1).value = "Filters applied:"
            self.summary[f"A{row_count}"].font = Font(
                bold=True, name=DEFAULT_FONT.name)
            self.summary.cell(row_count, 2).value = self.args.human_filter

            row_count += 2

        # write args passed to script to generate report
        self.summary.cell(row_count, 1).value = "Filter command:"
        self.summary[f"A{row_count}"].font = Font(bold=True, name=DEFAULT_FONT.name)
        if self.args.filter:
            self.summary.cell(row_count, 2).value = self.args.filter
        else:
            self.summary.cell(row_count, 2).value = "None"

        row_count += 2

        # write in the colouring of any columns if done
        if self.args.colour:
            row_count, to_bold = self.summary_sheet_cell_colour_key(
                row_count, to_bold)

        # write more text with DNAnexus IDs etc
        row_count += 2
        self.summary.cell(row_count, 1).value = "Workflow:"
        self.summary.cell(row_count + 1, 1).value = "Workflow ID:"
        self.summary.cell(row_count + 2, 1).value = "Report Job ID:"
        to_bold.extend([f"A{row_count + x}" for x in range(0, 3)])

        self.summary.cell(row_count, 2).value = self.args.workflow[0]
        self.summary.cell(row_count + 1, 2).value = self.args.workflow[1]
        self.summary.cell(row_count + 2, 2).value = self.args.job_id

        for cell in to_bold:
            self.summary[cell].font = Font(bold=True, name=DEFAULT_FONT.name)


    def dias_summary(self) -> None:
        """
        Write summary sheet in format for RD group, adds the following info:
            - sample ID, panel(s), run IDs etc.
            - formatted tables for them to fill in reporting
        """
        details_dict = defaultdict()
        # write titles for summary values
        self.summary.cell(1, 1).value = "Sample ID:"
        self.summary.cell(1, 5).value = "Clinical Indication(s):"
        self.summary.cell(2, 5).value = "Panel(s):"
        self.summary.cell(33, 1).value = "Total records:"

        # get sample name from vcf, should only be one but handle everything
        # list-wise just in case
        sample = [
            Path(x).name.replace('.vcf', '').replace('.gz', '')
            for x in self.args.vcfs
        ]
        sample = [x.split('_')[0] if '_' in x else x for x in sample]
        sample = str(sample).strip('[]').strip("'")

        # write summary values
        self.summary.cell(1, 2).value = sample
        self.summary.cell(1, 6).value = self.args.clinical_indication
        self.summary.cell(2, 6).value = self.args.panel

        # If clinical indication given as arg, add this to our dict
        # Write out the dias clinical indication info to JSON file
        if self.args.clinical_indication:
            details_dict['clinical_indication'] = self.args.clinical_indication
            with open('details.json', 'w', encoding='utf8') as details_json:
                json.dump(details_dict, details_json)

        # write total rows in each sheet
        count = 33

        # cells to make bold
        to_bold = []

        for sheet, vcf in zip(self.args.sheets, self.vcfs):
            self.summary.cell(count, 2).value = sheet
            self.summary.cell(count, 3).value = len(vcf.index)
            to_bold.append(f"A{count}")
            count += 1


        count += 5

        # write genome reference(s) parsed from vcf header
        if self.refs:
            self.summary.cell(count, 1).value = "Reference:"
            self.summary[f"A{count}"].font = Font(
                bold=True, name=DEFAULT_FONT.name
            )
            for ref in list(set(self.refs)):
                self.summary.cell(count, 2).value = ref
                count += 1

        count += 1

        if self.args.human_filter:
            self.summary.cell(count, 1).value = "Filters applied:"
            self.summary.cell(count, 2).value = self.args.human_filter
            to_bold.append(f"A{count}")

            count += 2

        # write args passed to script to generate report
        self.summary.cell(count, 1).value = "Filter command:"
        to_bold.append(f"A{count}")
        if self.args.filter:
            self.summary.cell(count, 2).value = self.args.filter
        else:
            self.summary.cell(count, 2).value = "None"

        count += 2

        if self.args.colour:
            count, to_bold = self.summary_sheet_cell_colour_key(
                count, to_bold)

        count += 2

        self.summary.cell(count, 1).value = "Workflow:"
        self.summary.cell(count + 1, 1).value = "Workflow ID:"
        self.summary.cell(count + 2, 1).value = "Report Job ID:"
        to_bold.append(f"A{count}")
        to_bold.append(f"A{count + 1}")
        to_bold.append(f"A{count + 2}")

        self.summary.cell(count, 2).value = self.args.workflow[0]
        self.summary.cell(count + 1, 2).value = self.args.workflow[1]
        self.summary.cell(count + 2, 2).value = self.args.job_id


        # write center reporting section tables
        self.summary.cell(2, 1).value = "Lab no."
        self.summary.cell(2, 3).value = "First name"
        self.summary.cell(2, 4).value = "Last name"
        self.summary.cell(4, 1).value = "Number checked"
        self.summary.cell(5, 1).value = "Summary coverage"
        self.summary.cell(14, 2).value = "Phenotype:"

        self.summary.cell(21, 2).value = "Panels"
        self.summary.cell(21, 3).value = "Excel file"
        self.summary.cell(21, 4).value = "Comments"
        self.summary.cell(21, 6).value = "Analysis by"
        self.summary.cell(21, 7).value = "Date"
        self.summary.cell(21, 8).value = "Checked by"
        self.summary.cell(21, 9).value = "Date"

        self.summary.cell(26, 2).value = "Sanger sequencing confirmation"
        self.summary.cell(27, 2).value = "Gene"
        self.summary.cell(27, 3).value = "NM_#"
        self.summary.cell(27, 4).value = "Coordinate"
        self.summary.cell(27, 5).value = "cDNA"
        self.summary.cell(27, 6).value = "Protein change"
        self.summary.cell(27, 7).value = "WS#"
        self.summary.cell(27, 8).value = "Confirmed (Y/N)"

        # merge some title columns that have longer text
        self.summary.merge_cells(
            start_row=1, end_row=1, start_column=2, end_column=4)
        self.summary.merge_cells(
            start_row=1, end_row=1, start_column=6, end_column=20)
        self.summary.merge_cells(
            start_row=2, end_row=2, start_column=6, end_column=20)
        self.summary.merge_cells(
            start_row=14, end_row=14, start_column=2, end_column=5)
        self.summary.merge_cells(
            start_row=26, end_row=26, start_column=2, end_column=8)
        self.summary.merge_cells(
            start_row=21, end_row=21, start_column=4, end_column=5)
        self.summary.merge_cells(
            start_row=5, end_row=11, start_column=1, end_column=1)
        
        # make the coverage tile centre of merged rows
        self.summary["A5"].alignment = Alignment(
                           wrapText=True, vertical="center")

        # titles to set to bold
        to_bold += [
                "A1", "A2", "A4", "A5", "A33", "B1", "B14", "B21", "B26",
                "B27", "B33", "B34", "C2", "C21", "C27", "D2", "D21", "D27",
                "E1", "E2", "E27", "F21", "F27", "G21", "G27", "H21", "H27",
                "I21"
                ]

        for cell in to_bold:
            self.summary[cell].font = Font(bold=True, name=DEFAULT_FONT.name)

        # set column widths for readability
        self.summary.column_dimensions['A'].width = 22 if self.args.colour else 18
        self.summary.column_dimensions['B'].width = 13
        self.summary.column_dimensions['C'].width = 13
        self.summary.column_dimensions['D'].width = 13
        self.summary.column_dimensions['E'].width = 22
        self.summary.column_dimensions['F'].width = 16
        self.summary.column_dimensions['G'].width = 16
        self.summary.column_dimensions['H'].width = 16

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="0CABA8")

        colour_cells = [
            "A2", "A4", "A5", "B2", "B14", "B21", "B26", "B27",
            "C2", "C21", "C27", "D2", "D21", "D27", "E27", "F21",
            "F27", "G21", "G27", "H21", "H27", "I21"
        ]
        for cell in colour_cells:
            self.summary[cell].fill = blueFill

        # set borders around table areas
        row_ranges = [
            'A2:D2', 'A3:D3', 'A4:D4', 'A5:A5', 'B14:E14', 'B15:E15',
            'B16:E16', 'B17:E17', 'B18:E18', 'B21:I21', 'B22:I22',
            'B23:I23', 'B26:H26', 'B27:H27', 'B28:H28', 'B29:H29', 'B30:H30'
            ]
        for row in row_ranges:
            for cells in self.summary[row]:
                for cell in cells:
                    cell.border = THIN_BORDER
        if self.args.lock_sheet:
            cell_to_unlock = ["A3", "B3", "B4", "B5", "B6", "B7", "B8", "B9",
                              "B10", "B11", "C3", "C4", "C5", "C6", "C7", "C8",
                              "C9", "C10", "C11", "D3", "D4", "D5", "D6", "D7",
                              "D8", "D9", "D10", "D11", "B15", "C15", "D15",
                              "E15", "B16", "C16", "D16", "E16", "B17", "C17",
                              "D17", "E17", "B18", "C18", "D18", "E18", "B22",
                              "C22", "D22", "E22", "F22", "G22", "H22", "I22",
                              "B23", "C23", "D23", "E23", "F23", "G23", "H23",
                              "I23", "B28", "C28", "D28", "E28", "F28", "G28",
                              "H28", "B29", "C29", "D29", "E29", "F29", "G29",
                              "H29", "B30", "C30", "D30", "E30", "F30", "G30",
                              "H30"
                              ]
            self.lock_sheet(ws=self.summary,
                            cell_to_unlock=cell_to_unlock,
                            start_row=self.summary.max_row+1,
                            start_col=10,
                            unlock_row_num=ROW_TO_UNLOCK,
                            unlock_col_num=COL_TO_UNLOCK)

    def write_reporting_template(self, report_sheet_num) -> None:
        """
        Writes sheet(s) to Excel file with formatting for reporting against
        ACMG criteria
        """
        report = self.workbook.create_sheet(f"interpret_{report_sheet_num}")

        titles = {
            "Gene": [2, 2],
            "HGVSc": [2, 3],
            "HGVSp": [2, 4],
            "EVIDENCE": [8, 3],
            "PATHOGENIC": [8, 7],
            "P_STRENGTH": [8, 8],
            "P_POINTS": [8, 9],
            "BENIGN": [8, 10],
            "B_STRENGTH": [8, 11],
            "B_POINTS": [8, 12],
            "Associated disease": [4, 2],
            "Known inheritance": [5, 2],
            "Prevalence": [6, 2],
            ("Allele frequency is >5% (or gene-specific cut off) in "
             "population data e.g. gnomAD, UKB"): [9, 2],
            ("Null variant in a gene where LOF is known mechanism "
             "of disease\nand non-canonical splice variants where "
             "RNA analysis confirms\naberrant transcription"): [10, 2],
            ("Same AA change as previously established pathogenic "
             "variant\nregardless of nucleotide change and splicing "
             "variants within\nsame motif with identical predicted "
             "effect"): [11, 2],
            ("De novo (confirmed) / observed in\nhealthy adult "
             "with full penetrance expected at an early age"): [12, 2],
            "In vivo / in vitro functional studies": [13, 2],
            "Prevalence in affected > controls": [14, 2],
            ("In mutational hot spot and/or critical functional "
             "domain, without\nbenign variation"): [15, 2],
            ("Freq in controls eg gnomAD, low/absent (PM2) or allele "
             "frequency is greater than expected for disorder (BS1)"): [16, 2],
            "Detected in trans/in cis with pathogenic variant": [17, 2],
            ("In frame protein length change/stop-loss variants, "
             "non repeat\nvs. repeat region"): [18, 2],
            ("Missense change at AA where different likely/pathogenic\n"
             "missense change seen before"): [19, 2],
            "Assumed de novo (no confirmation)": [20, 2],
            "Cosegregation with disease in family, not in unaffected": [21, 2],
            ("Missense where low rate of benign missense and common\n"
             "mechanism (Z score ≥3.09), or missense where LOF common\n"
             "mechanism"): [22, 2],
            "Multiple lines of computational evidence": [23, 2],
            ("Phenotype/FH specific for disease of single etiology, or\n"
             "alternative genetic cause of disease detected"): [24, 2],
            ("Synonymous change, no affect on splicing, not conserved; "
             "splice\nvariants confirmed to have no impact"): [25, 2],
            "POINTS": [26, 7]
        }
        for key, val in titles.items():
            report.cell(val[0], val[1]).value = key
            report.cell(val[0], val[1]).font = Font(
                bold=True, name=DEFAULT_FONT.name
            )
        classifications = {
            "PVS1": [(10, 7)],
            "PS1": [(11, 7)],
            "PS2": [(12, 7)],
            "PS3": [(13, 7)],
            "PS4": [(14, 7)],
            "PM1": [(15, 7)],
            "PM2": [(16, 7)],
            "PM3": [(17, 7)],
            "PM4": [(18, 7)],
            "PM5": [(19, 7)],
            "PM6": [(20, 7)],
            "PP1": [(21, 7)],
            "PP2": [(22, 7)],
            "PP3": [(23, 7)],
            "PP4": [(24, 7)],
            "BA1": [(9, 10)],
            "BS2": [(12, 10)],
            "BS3": [(13, 10)],
            "BS1": [(16, 10)],
            "BP2": [(17, 10)],
            "BP3": [(18, 10)],
            "BS4": [(21, 10)],
            "BP1": [(22, 10)],
            "BP4": [(23, 10)],
            "BP5": [(24, 10)],
            "BP7": [(25, 10)]
        }

        for key, values in classifications.items():
            for val in values:
                report.cell(val[0], val[1]).value = key

        # nice formatting of title text and columns
        for col in (['B', 'C', 'G', 'H', 'I', 'J', 'K', 'L']):
            for row in range(8, 26):
                if row == 8:
                    report[f"{col}{row}"].alignment = Alignment(
                           wrapText=True, vertical="center",
                           horizontal="center"
                    )
                elif (col == "B" and row != 8) or (col == "C" and row != 8):
                    report.row_dimensions[row].height = 50
                    report[f"{col}{row}"].alignment = Alignment(
                           wrapText=True, vertical="center"
                    )
                else:
                    report[f"{col}{row}"].alignment = Alignment(
                           wrapText=True, vertical="center",
                           horizontal="center"
                    )
                    report[f"{col}{row}"].font = Font(size=14,
                                                      name=DEFAULT_FONT.name)

        for col in (['B', 'C', 'D']):
            for row in range(2, 7):
                report.row_dimensions[row].height = 20
                report[f"{col}{row}"].font = Font(size=14, bold=True,
                                                  name=DEFAULT_FONT.name)

        # merge associated disease, inheritance and prevalence cells
        for row in range(4, 8):
            report.merge_cells(
                start_row=row, end_row=row, start_column=3, end_column=12)

        # merge evidence cells
        for row in range(8, 27):
            report.merge_cells(
                start_row=row, end_row=row, start_column=3, end_column=6)

        # merge POINTS cells
        report.merge_cells(
                start_row=26, end_row=26, start_column=8, end_column=12)

        # set appropriate widths
        report.column_dimensions['B'].width = 62
        report.column_dimensions['C'].width = 35
        report.column_dimensions['D'].width = 35
        report.column_dimensions['E'].width = 5
        report.column_dimensions['F'].width = 5
        report.column_dimensions['G'].width = 14
        report.column_dimensions['H'].width = 14
        report.column_dimensions['I'].width = 14
        report.column_dimensions['J'].width = 14
        report.column_dimensions['K'].width = 14
        report.column_dimensions['L'].width = 14

        # do some colouring
        colour_cells = {
            'E46C0A': ['G11', 'G12', 'G13', 'G14'],
            'FFC000': ['G15', 'G16', 'G17', 'G18', 'G19', 'G20'],
            'FFFF00': ['G21', 'G22', 'G23', 'G24'],
            '00B0F0': ['J12', 'J13', 'J16', 'J21'],
            '92D050': ['J17', 'J18', 'J22', 'J23', 'J24', 'J25'],
            '0070C0': ['J9'],
            'FF0000': ['G10'],
            'D9D9D9': ['G9', 'G25', 'H9', 'H25', 'I9', 'I25',
                       'J10', 'J11', 'J14', 'J15', 'J19', 'J20',
                       'K10', 'K11', 'K14', 'K15', 'K19', 'K20',
                       'L10', 'L11', 'L14', 'L15', 'L19', 'L20'
                       ]

        }
        for colour, cells in colour_cells.items():
            for cell in cells:
                report[cell].fill = PatternFill(
                    patternType="solid", start_color=colour
                )

        # add some borders
        row_ranges = {
            'horizontal': [
                'B3:D3', 'B4:J4', 'B5:L5',
                'B6:L6', 'B7:L7', 'B8:L8', 'B9:L9', 'B10:L10', 'B11:L11',
                'B12:L12', 'B13:L13', 'B14:L14', 'B15:L15', 'B16:L16',
                'B17:L17', 'B18:L18', 'B19:L19', 'B20:L20', 'B21:L21',
                'B22:L22', 'B23:L23', 'B24:L24', 'B25:L25'
            ],
            'horizontal_thick': [
                'B2:D2', 'B4:L4', 'B7:L7', 'B8:L8', 'B26:L26', 'B27:L27'
            ],
            'vertical': [
                'E2:E3', 'G8:G26', 'H8:H25', 'I8:I25', 'J8:J25',
                'K8:K25', 'L8:L25',
            ],
            'vertical_thick': [
                'B2:B6', 'B8:B26', 'C2:C6', 'C8:C26', 'M4:M6', 'M8:M26',
                'E2:E3'
            ]
        }

        for side, values in row_ranges.items():
            for row in values:
                for cells in report[row]:
                    for cell in cells:
                        # border style is immutable => copy current and modify
                        cell_border = cell.border.copy()
                        if side == 'horizontal':
                            cell_border.top = THIN
                        if side == 'horizontal_thick':
                            cell_border.top = MEDIUM
                        if side == 'vertical':
                            cell_border.left = THIN
                        if side == 'vertical_thick':
                            cell_border.left = MEDIUM
                        cell.border = cell_border
        if self.args.lock_sheet:
            cell_to_unlock = ["B3", "C3", "D3", "C4", "C5", "C6",
                              "C9", "C10", "C11", "C12", "C13", "C14", "C15",
                              "C16", "C17", "C18", "C19", "C20", "C21", "C22",
                              "C23", "C24", "C25", "C26", "H10", "H11",
                              "H12", "H13", "H14", "H15", "H16", "H17", "H18",
                              "H19", "H20", "H21", "H22", "H23", "H24", "I10",
                              "I11", "I12", "I13", "I14", "I15", "I16", "I17",
                              "I18", "I19", "I20", "I21", "I22", "I23", "I24",
                              "K9", "K12", "K13", "K16", "K17", "K18", "K21",
                              "K22", "K23", "K24", "K25", "L9", "L12", "L13",
                              "L16", "L17", "L18", "L21", "L22", "L23", "L24",
                              "L25", "H26"]
            self.lock_sheet(ws=report,
                            cell_to_unlock=cell_to_unlock,
                            start_row=report.max_row,
                            start_col=report.max_column,
                            unlock_row_num=ROW_TO_UNLOCK,
                            unlock_col_num=COL_TO_UNLOCK)

    def write_variants(self) -> None:
        """
        Writes all variants from dataframe(s) to sheet(s) specified in
        self.args.sheets.

        If sheet names not specified, these will be set as "variant" where one
        dataframe is being written, or the vcf filename prefix if there are
        more than one dataframes to write
        """
        total_rows = sum([len(x) for x in self.vcfs])
        print(f"\nWriting total of {total_rows} rows to output xlsx file")
        if total_rows > 5000:
            print(
                "Writing many rows to Excel is slow, "
                "this may take a few minutes..."
            )

        # If details.json already exists (dias summary page written and
        # clinical indication is given as arg) then
        # open it and read it in so we can add var counts to the dict
        # otherwise just make a new empty dict
        if os.path.isfile('details.json'):
            with open('details.json', 'r', encoding='utf8') as details_json:
                details_dict = json.load(details_json)
        else:
            details_dict = defaultdict()

        with self.writer:
            # add variants
            for sheet, vcf in zip(self.args.sheets, self.vcfs):
                sheet_no = self.args.sheets.index(sheet) + 1
                print(
                    f"\nWriting {len(vcf)} rows to {sheet} sheet "
                    f"({sheet_no}/{len(self.args.sheets)})"
                )
                details_dict[sheet] = len(vcf)

                # timing how long it takes to write because its slow
                start = timer()
                vcf.to_excel(
                    self.writer, sheet_name=sheet,
                    index=False, float_format="%.3f"
                )

                curr_worksheet = self.writer.sheets[sheet]
                self.set_widths(curr_worksheet, vcf.columns)
                self.set_font(curr_worksheet)
                self.colour_hyperlinks(curr_worksheet)
                self.colour_cells(curr_worksheet)
                self.set_dp(curr_worksheet)

                # freeze header so scrolling keeps it in view
                curr_worksheet.freeze_panes = self.args.freeze_column

                self.workbook.save(self.args.output)
                end = timer()
                print(
                    f"Writing to Excel for sheet {sheet} took "
                    f"{round(end - start)}s"
                )

                # read back the written sheet to check its written correctly
                self.check_written_sheets(vcf, sheet)

                # set Excel types for numeric cells to suppress Excel warnings
                self.set_types(curr_worksheet)
                if self.args.acmg and self.args.lock_sheet:
                    num_variant = vcf.shape[0]
                    cell_to_unlock = []
                    comment_col = self.get_col_letter(curr_worksheet,
                                                      "Comment")
                    interpreted_col = self.get_col_letter(curr_worksheet,
                                                          "Interpreted")
                    for row in range(2, num_variant+2):
                        if comment_col is not None:
                            cell_to_unlock.append(f"{comment_col}{row}")
                        if curr_worksheet.title == self.args.sheets[0]:
                            cell_to_unlock.append(f"{interpreted_col}{row}")
                    self.lock_sheet(ws=curr_worksheet,
                                    cell_to_unlock=cell_to_unlock,
                                    start_row=num_variant+2,
                                    start_col=curr_worksheet.max_column+1,
                                    unlock_row_num=ROW_TO_UNLOCK,
                                    unlock_col_num=COL_TO_UNLOCK)
                self.workbook.save(self.args.output)

        # Write out dict to file
        with open('details.json', 'w', encoding='utf8') as details_json:
            json.dump(details_dict, details_json)


    def write_additional_files(self) -> None:
        """
        Write each dataframe of additional files passed to separate sheets
        """
        if not self.additional_files:
            # empty dict => no files passed to write
            return

        print("Writing additional file(s) to workbook")

        for file_name, file_df in self.additional_files.items():
            file_df.to_excel(
                self.writer, sheet_name=file_name, index=False, header=None
            )

            curr_worksheet = self.writer.sheets[file_name]
            self.set_font(curr_worksheet)
            self.set_types(curr_worksheet)

            # set appropriate column widths based on cell contents
            for idx, column in enumerate(curr_worksheet.columns, start=1):
                # get max length of column contents, sensible max and min sizes
                length = max(len(str(cell.value)) for cell in column)
                length = 13 if length < 13 else length
                length = 30 if length > 30 else length

                col_letter = get_column_letter(idx)
                curr_worksheet.column_dimensions[col_letter].width = length

            # set widths of any columns we have specified below in set_width()
            # to override the above defaults
            # get column names from first row of sheet
            sheet_columns = [
                '' if x is None else x for x in file_df.iloc[0].tolist()]
            self.set_widths(curr_worksheet, sheet_columns)

            if file_df.iloc[0].iloc[0] == 'Metric (UOM)':
                # additional file is MetricsOutput.tsv from TSO500 => attempt
                # to colour metrics in output sheet
                if len(file_df.columns.tolist()) == 4:
                    # only 4 columns => given sample metrics correctly
                    # parsed from full run metrics
                    try:
                        self.colour_metrics_output(file_df, curr_worksheet)
                    except Exception as err:
                        # catch any error raised to not break the app and
                        # just print a warning since its non-essential
                        print(
                            "Warning: error in colouring MetricsOutput sheet:"
                            f"\n\t{err}\nContinuing without colouring."
                        )


    def write_images(self) -> None:
        """
        Writes each of the passed images to a separate sheet
        """
        if not self.args.images:
            # no images to write
            return

        print("Writing image(s) to workbook")

        for idx, image in enumerate(self.args.images):
            if self.args.image_sheets:
                # names for image sheets specified
                sheet = self.workbook.create_sheet(
                    self.args.image_sheets[idx])
            else:
                sheet = self.workbook.create_sheet(f'image_{idx + 1}')

            img = drawing.image.Image(image)
            img.anchor = 'B2'

            if self.args.image_sizes:
                # set sizes if specified
                try:
                    width, height = self.args.image_sizes[idx].split(':')
                    img.height = float(height)
                    img.width = float(width)
                except Exception as error:
                    print(
                        "Failed to parse width/height from image_sizes "
                        f"argument for image no. {idx}, sizes specified: "
                        f"{self.args.image_sizes}.\n\nError: {error}\n\n"
                        f"Will use defaults from current image (width:"
                        f"{img.width}px, height:{img.height}px)"
                    )
            else:
                # set max size based off aspect ratio of original image
                ratio = img.width / img.height
                height = 972
                width = 972 * ratio

                img.height = height
                img.width = width

            sheet.add_image(img)


    def check_written_sheets(self, vcf, sheet) -> None:
        """"
        Check that the written sheet is exactly the same as the dataframe
        that was meant to be written

        Parameters
        ----------
        vcf : pd.DataFrame
            dataframe of sheet that was written to file
        sheet : str
            sheet name to read from file

        Raises
        ------
        AssertionError
            Raised when data written to sheet does not match the dataframe
        """
        print(f"\nVerifying data written to file for {sheet} sheet...\n")

        # read in written sheet using openpyxl to deal with Excel oddities
        sheet_data = load_workbook(filename=self.args.output)[sheet]
        written_sheet = pd.DataFrame(
            sheet_data.values, columns=vcf.columns.tolist())
        written_sheet = written_sheet.iloc[1:]  # drop header on first row

        # force nan values to be None strings for consistency on comparing
        vcf = vcf.replace(r'^\s*$', np.nan, regex=True)
        vcf.fillna('None', inplace=True)
        written_sheet = written_sheet.replace(r'^\s*$', np.nan, regex=True)
        written_sheet.fillna('None', inplace=True)

        # set all columns of both dfs to strings
        vcf = vcf.astype(str)
        written_sheet = written_sheet.astype(str).reset_index(drop=True)

        # floats with trailing zero seem inconsistent when writing to file,
        # since we've cast everything to a string strip for ease of comparing
        vcf = vcf.applymap(
            lambda x: x.replace('.0', '') if x.endswith('.0') else x
        )
        written_sheet = written_sheet.applymap(
            lambda x: x.replace('.0', '') if x.endswith('.0') else x
        )

        assert vcf.equals(written_sheet), (
            f"Written data for sheet: {sheet} does not seem to match the "
            "dataframe to be written"
        )


    def set_types(self, worksheet) -> None:
        """
        Iterate over all worksheet cells and test if cell value can be numeric,
        if so sets the type to numeric to suppress 'Number stored as text'
        warning in Excel

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        """
        for cells in worksheet.rows:
            for cell in cells:
                if is_numeric(cell.value):
                    cell.data_type = 'n'


    def set_font(self, worksheet) -> None:
        """
        Set font to all cells in variant sheet to Calibri

        Default is Times New Roman and it is ugly

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        """
        for cells in worksheet.rows:
            for cell in cells:
                cell.font = Font(name=DEFAULT_FONT.name)


    def set_dp(self, worksheet) -> None:
        """
        Set the dp to display values to (i.e. 0.14236 to dp -> 0.14).

        Original value will remain unchanged in the formula bar
        on selecting the cell, just the view of data is adjusted
        through Excel number formatting


        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        """
        # mapping of column names to no. dp to display
        col_to_dp = {
            'VF': 2
        }

        for column, dp in col_to_dp.items():
            dp = '0' * dp
            for ws_column in worksheet.iter_cols(1, worksheet.max_column):
                if ws_column[0].value.lower() == column.lower():
                    # column is a match, loop over every cell in the column
                    # to set formatting since there's no nicer way to apply
                    # the style in openpyxl
                    for row in ws_column:
                        row.number_format = f'#,##0.{dp}'


    def convert_colour(self, colour) -> str:
        """
        Converts string of colour to aRGB value that openpyxl will accept.

        Valid colour strings reference here:
        https://github.com/vaab/colour/blob/11f138eb7841d2045160b378a2eec0c2321144c0/colour.py#L52

        Parameters
        ----------
        colour : str
            colour string, either hex value beginning with '#' (#82920c)
            or human readable (ForestGreen)

        Returns
        -------
        str
            hex value without '#' prefix
        """
        if colour.startswith('#'):
            colour = colour.lstrip('#')
        else:
            # not given as '#FAC090' => assume its given as 'green|red' etc
            colour = Color(colour)
            # colour.saturation = 0.8
            colour = colour.hex_l.lstrip('#')

        return colour


    def colour_hyperlinks(self, worksheet) -> None:
        """
        Set text colour to blue if text contains hyperlink

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        """
        for cells in worksheet.rows:
            for cell in cells:
                if 'HYPERLINK' in str(cell.value):
                    cell.font = Font(color='00007f', name=DEFAULT_FONT.name)


    def colour_cells(self, worksheet) -> None:
        """
        Conditionally colours cells in a variant worksheet by user specified
        conditions and colours for a given column with --colour.

        Arguments will be in the following formats:
            - VF:>=0.9:green        => single condition
            - VF:<0.8&>=0.6:orange  => both of two conditions
            - VF:>0.9|<0.1:red      => either of two conditions

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet

        Raises
        ------
        ValueError
            Raised when invalid parameter passed
        """
        if not self.args.colour:
            # no cell colours defined
            return

        # mapping table of valid operators to operator methods
        ops = {
            "=": operator.eq,
            "!=": operator.ne,
            "+": operator.add,
            "-": operator.sub,
            ">": operator.gt,
            ">=": operator.ge,
            "<": operator.lt,
            "<=": operator.le
        }

        # dict to add any previously coloured cells that are likely from
        # overlapping expressions -> raise error if anything is present
        errors = defaultdict(dict)

        for column_to_colour in self.args.colour:
            column, conditions, colour = column_to_colour.split(':')
            column = column.replace('CSQ_', '')

            _and = False
            _or = False

            # check if more than one condition is passed and how to interpret
            if '&' in conditions:
                _and = True
                conditions = conditions.split('&')
            elif '|' in conditions:
                _or = True
                conditions = conditions.split('|')
            else:
                conditions = [conditions]

            colour = self.convert_colour(colour)

            # list of tuples to build as (operator, value)
            conditions_list = []

            # split out each operator and value to a list of tuples
            for condition in conditions:
                current_operator = re.match(r'(>=|<=|>|<|=|!=|\+|-)', condition)

                if not current_operator:
                    # invalid or no operator passed
                    raise ValueError(
                        "Invalid operator passed for cell colouring in "
                        f"argument: {column_to_colour}"
                    )

                # add to list of tuples of operators and corresponding value
                current_operator = current_operator.group()
                value = condition.replace(current_operator, '')
                if is_numeric(value):
                    value = float(value)

                conditions_list.append((current_operator, value))

            # find correct column to colour, then colour cells according
            # to the given conditions
            for column_cells in worksheet.iter_cols(1, worksheet.max_column):
                if column_cells[0].value == column:
                    for cell in column_cells[1:]:
                        # first test if cell value is numeric for comparing
                        if is_numeric(cell.value):
                            cell_value = float(cell.value)
                        else:
                            cell_value = cell.value

                        if _and:
                            to_colour = all([
                                True if ops[condition[0]](cell_value, condition[1])
                                else False for condition in conditions_list
                                ])
                        elif _or:
                            to_colour = any([
                                True if ops[condition[0]](cell_value, condition[1])
                                else False for condition in conditions_list
                                ])
                        else:
                            # should just be one condition as no & or |
                            current_operator, value = conditions_list[0]
                            to_colour = ops[current_operator](cell_value, value)

                        if to_colour:
                            cell_colour = cell.fill.start_color.index
                            if not cell_colour == '00000000':
                                # cell already coloured => add to errors
                                warn = errors.get((cell_colour, colour), [])
                                warn.append(cell.coordinate)
                                errors[(cell_colour, colour)] = warn
                            else:
                                worksheet[cell.coordinate].fill = PatternFill(
                                    patternType="solid",
                                    start_color=colour
                                )
        if errors:
            error_message = (
                f"\n{'#' * 35} ERROR {'#' * 35}\n\n"
                "Overlapping colouring of cells, the following "
                "cells colour were not changed due \nto being previously coloured:"
            )
            for colours, cells in errors.items():
                if len(cells) > 5:
                    cell_count = len(cells) - 5
                    cells = f"{', '.join(cells[:5])} + {cell_count} more cells"

                error_message += (
                    f"\n\tCurrent cell colour: {colours[0]}"
                    f"\n\tNew cell colour: {colours[1]}"
                    f"\n\tCells not coloured: {cells}\n"
                )

            error_message += f"\n{'#' * 79}"

            raise RuntimeError(error_message)


    def set_widths(self, worksheet, sheet_columns) -> None:
        """
        Set widths for variant sheets off common names to be more readable,
        calls get_closest_match() method to determine appropriate column
        width to use.

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        sheet_columns : list
            column names for sheet from DataFrame.columns
        """
        widths = {
            "chrom": 7,
            "pos": 12,
            "ref": 10,
            "alt": 10,
            "qual": 10,
            "dp": 10,
            'ac': 10,
            'af': 10,
            'an': 10,
            'baseqranksum': 15,
            'clippingranksum': 16,
            "symbol": 9,
            "exon": 9,
            "variant class": 15,
            "consequence": 17,
            "hgvsc": 27,
            "hgvsp": 27,
            "hgvsg": 18,
            "dna": 12,
            "protein": 13,
            "gnomad_af": 16,
            "gnomad_exomes_af": 16,
            "gnomad_genomes_af": 16,
            "hgmd": 13,
            "hgmd_phen": 15,
            "hgmd_rankscore": 16,
            "spliceai_ds": 18,
            "existing variation": 18,
            "clinvar": 10,
            "clinvar clndn": 18,
            "clinvar clinsig": 18,
            "cosmic": 15,
            "feature": 13,
            "decipher": 24,
            "Metric (UOM)": 52,  # TSO500 MetricsOutput.tsv
            "[TMB]": 32,  # TSO500 CombinedVariantOutput.tsv
            "rawchange": 20,
            "vf":6,
            "comment": 10,
            "classification": 12,
            "spliceai pred ": 18
        }

        # generate list of 286 potential xlsx columns from A,B,C...JX,JY,JZ
        # allows for handling a lot of columns
        column_list = [
            f"{x}{y}" for x in [
                '', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'
            ] for y in uppercase
        ]

        for idx, column in enumerate(sheet_columns):
            # loop over column names, select width by closest match from dict
            # and set width in sheet letter
            width = self.get_closest_match(
                worksheet, column_list[idx], column.lower(), widths
            )
            worksheet.column_dimensions[column_list[idx]].width = width


    def get_closest_match(self, worksheet, col_letter, col, widths) -> int:
        """
        Given a column name, find the closest match (if there is one) in the
        widths dict and return its width value to set. Using imprecise name
        matching as columns can differ between variant callers, annotation in
        VEP and renaming in vcf.rename_columns()

        Parameters
        ----------
        worksheet : openpyxl.Writer
            writer object for current sheet
        col_letter : str
            letter of column in worksheet
        col : str
            name of column
        widths : dict
            dict of common column names and widths

        Returns
        -------
        width : int
            column width value to set
        """
        distances = {x: levenshtein.distance(col, x) for x in widths.keys()}
        closest_match = min(distances, key=distances.get)

        if distances[closest_match] <= 5:
            # close enough match to probably be correct
            width = widths[closest_match]
        else:
            # no close matches to name, use title multipled by factor
            title = worksheet[f"{col_letter}1"].value
            width = len(title) * 1.15
            if width < 13:
                # make minimum of 13
                width = 13

        return width


    def colour_metrics_output(self, file_df, worksheet) -> None:
        """
        Add colouring to MetricsOutput sheet, this will colour a defined no.
        of rows dependent on the sample value and upper and lower limits

        File is formatted as:

        Metric (UOM)                LSL Guideline	USL Guideline	Sample
        COVERAGE_MAD (Count)        0	            0.21	        0.12
        MEDIAN_BIN_COUNT_CNV_TARGET	1	            NA	            6.1


        Where in the above, both samples values would be coloured green as
        COVERAGE_MAD lies between LSL and USL, and MEDIAN_BIN_COUNT_CNV_TARGET
        is above the LSL.


        Parameters
        ----------
        file_df : pd.DataFrame
            DataFrame of MetricsOutput written to the workbook sheet
        worksheet : openpyxl.Writer
            writer object for current sheet
        """
        to_colour = [
            2, 6, 7, 12, 16, 17, 21, 22, 23, 24, 26, 27, 28, 29,
            30, 31, 32, 33, 34, 38, 39, 40, 44, 45, 46, 47
        ]
        green = []
        amber = []
        red = []

        for idx, row in file_df.iterrows():
            if not all([row.iloc[1], row.iloc[2], row.iloc[3]]):
                # blank row
                continue
            if idx in to_colour:
                if row.iloc[1] == 'NA' and row.iloc[2] == 'NA':
                    # both have no guideline values => skip
                    continue
                if row.iloc[3] == 'NA':
                    # no sample value
                    continue
                if row.iloc[1] != 'NA' and row.iloc[2] == 'NA':
                    # lower limit but no upper limit
                    if float(row.iloc[3]) >= float(row.iloc[1]):
                        green.append(idx)
                    else:
                        red.append(idx)
                if row.iloc[1] == 'NA' and row.iloc[2] != 'NA':
                    # no lower limit but has upper limit
                    if float(row.iloc[3]) <= float(row.iloc[2]):
                        green.append(idx)
                    else:
                        red.append(idx)
                if row.iloc[1] != 'NA' and row.iloc[2] != 'NA':
                    # lower and upper limits set:
                    if float(row.iloc[1]) <= float(row.iloc[3]) <= float(row.iloc[2]):
                        green.append(idx)
                    else:
                        red.append(idx)

        # PCT EXON 50x and 100x using more stringent thresholds than in file
        if float(file_df.iloc[:, 3][8]) >= 95:
            # 50x
            green.append(8)
        else:
            red.append(8)

        if float(file_df.iloc[:, 3][25]) >= 90:
            # 100x
            green.append(25)
        else:
            red.append(25)

        # contamination score wants to be amber if over upper bound
        if float(file_df.iloc[:, 3][1]) > float(file_df.iloc[:, 2][1]):
            amber.append(1)
        else:
            green.append(1)

        to_colour.extend([1, 8, 25])

        for idx in to_colour:
            if idx in green:
                worksheet[f"D{idx+1}"].fill = PatternFill(
                    patternType="solid",
                    start_color='008100'
                )
            if idx in amber:
                worksheet[f"D{idx+1}"].fill = PatternFill(
                    patternType="solid",
                    start_color='ff9f00'
                )
            if idx in red:
                worksheet[f"D{idx+1}"].fill = PatternFill(
                    patternType="solid",
                    start_color='b30000'
                )

        # add explanation on colouring
        worksheet["F3"].value = (
            "Colouring in this sheet is based off the sample value being "
            "between the LSL and USL \nguidelines, with the following exceptions:"
        )
        worksheet["F5"].value = CellRichText("- PCT_EXON_50X LSL set to ",
            TextBlock(InlineFont(b=True, rFont='Calibri'), '95'))
        worksheet["F6"].value = CellRichText("- PCT_EXON_100X LSL set to ",
            TextBlock(InlineFont(b=True, rFont='Calibri'), '90'))
        worksheet["F7"].value = "- CONTAMINATION_SCORE > USL will be amber"

        worksheet.merge_cells(
            start_row=3, end_row=4, start_column=6, end_column=14)
        worksheet.merge_cells(
            start_row=5, end_row=5, start_column=6, end_column=10)
        worksheet.merge_cells(
            start_row=6, end_row=6, start_column=6, end_column=10)
        worksheet.merge_cells(
            start_row=7, end_row=7, start_column=6, end_column=10)

    def drop_down(self) -> None:
        """
        Function to add drop-downs in the report tab for entering
        ACMG criteria for classification, as well as a boolean
        drop down into the additional 'Interpreted' column of
        the variant sheet(s).
        """
        wb = load_workbook(filename=self.args.output)

        # adding dropdowns in report table
        for sheet_num in range(1, self.args.acmg+1):
            # adding strength dropdown except for BA1
            report_sheet = wb[f"interpret_{sheet_num}"]
            cells_for_strength = ['H10', 'H11', 'H12', 'H13', 'H14', 'H15',
                                  'H16', 'H17', 'H18', 'H19', 'H20', 'H21',
                                  'H22', 'H23', 'H24', 'K12', 'K13', 'K16',
                                  'K17', 'K18', 'K21', 'K22', 'K23', 'K24',
                                  'K25']
            strength_options = '"Very Strong, Strong, Moderate, \
                                 Supporting, NA"'
            self.get_drop_down(dropdown_options=strength_options,
                               prompt='Select from the list',
                               title='Strength',
                               sheet=report_sheet,
                               cells=cells_for_strength)

            # add stregth for BA1
            BA1_options = '"Stand-Alone, Very Strong, Strong, Moderate, \
                            Supporting, NA"'
            self.get_drop_down(dropdown_options=BA1_options,
                               prompt='Select from the list',
                               title='Strength',
                               sheet=report_sheet,
                               cells=['K9'])

            # adding final classification dropdown
            report_sheet['B26'] = 'FINAL ACMG CLASSIFICATION'
            report_sheet['B26'].font = Font(bold=True, name=DEFAULT_FONT.name)
            class_options = '"Pathogenic,Likely Pathogenic, \
                              Uncertain Significance, \
                              Likely Benign, Benign"'
            self.get_drop_down(dropdown_options=class_options,
                               prompt='Select from the list',
                               title='ACMG classification',
                               sheet=report_sheet,
                               cells=['C26'])

        # adding Interpreted column dropdown in the first variant sheet tab
        first_variant_sheet = wb[self.args.sheets[0]]
        interpreted_options = '"YES,NO"'
        col_letter = self.get_col_letter(first_variant_sheet, "Interpreted")
        num_variant = self.vcfs[0].shape[0]
        if num_variant > 0:
            cells_for_variant = []
            for i in range(num_variant):
                cells_for_variant.append(f"{col_letter}{i+2}")
            self.get_drop_down(dropdown_options=interpreted_options,
                               prompt='Choose YES or NO',
                               title='Variant interpreted or not?',
                               sheet=first_variant_sheet,
                               cells=cells_for_variant)
        wb.save(self.args.output)

    def lock_sheet(self, ws, cell_to_unlock, start_row, start_col,
                   unlock_row_num, unlock_col_num) -> None:
        """
        locking the workbooksheet (password protected) and unlocking
        specific cells inside the table and regions outside table

        Parameters:
        -----------
        ws: str
            current worksheet
        cell_to_unlock: list
            list containing cells to unlock
        start_row: int
            integer indicating row starting to unlock
        start_col: int
            integer indicating col starting to unlock
        unlock_row_num: int
            integer indication number of row(s) to unlock
        unlock_col_num: int
            integer indication number of col(s) to unlock
        """
        ws.protection.sheet = True
        ws.protection.password = "sheet_is_protected"

        # unlocking specific cells inside the table
        for cell in cell_to_unlock:
            ws[cell].protection = Protection(locked=False)

        # unlocking regions outside table
        for col in range(1, start_col+unlock_col_num):
            col_letter = get_column_letter(col)
            for row in range(start_row, start_row+unlock_row_num):
                row_num = row
                cell = f"{col_letter}{row_num}"
                ws[cell].protection = Protection(locked=False)
        for col in range(start_col, start_col+unlock_col_num):
            col_letter = get_column_letter(col)
            for row in range(1, start_row):
                row_num = row
                cell = f"{col_letter}{row_num}"
                ws[cell].protection = Protection(locked=False)
        prot = ws.protection
        prot.formatColumns = False
        prot.formatRows = False
        prot.formatCells = False

    def get_col_letter(self, worksheet, col_name) -> str:
        """
        Getting the column letter with specific col name

        Parameters
        ----------
        worksheet: openpyxl.Writer
               writer object of current sheet
        col_name: str
               name of column to get col letter
        Return
        -------
        str
            column letter for specific column name
        """
        col_letter = None
        for column_cell in worksheet.iter_cols(1, worksheet.max_column):
            if column_cell[0].value == col_name:
                col_letter = column_cell[0].column_letter

        return col_letter

    def protect_rename_sheets(self) -> None:
        """
        prevent renaming sheets in the workbook
        """
        wb = load_workbook(filename=self.args.output)
        wb.security.lockStructure = True
        wb.security.workbookPassword = "sheet_name_protected"
        wb.save(self.args.output)

    def get_drop_down(self, dropdown_options, prompt, title, sheet, cells) -> None:
        """
        create the drop-downs items for designated cells

        Parameters
        ----------
        dropdown_options: str
            str containing drop-down items
        prompt: str
            prompt message for drop-down
        title: str
            title message for drop-down
        sheet: openpyxl.Writer writer object
            current worksheet
        cells: list
            list of cells to add drop-down
        """
        options = dropdown_options
        val = DataValidation(type='list', formula1=options,
                             allow_blank=True)
        val.prompt = prompt
        val.promptTitle = title
        sheet.add_data_validation(val)
        for cell in cells:
            val.add(sheet[cell])
        val.showInputMessage = True
        val.showErrorMessage = True

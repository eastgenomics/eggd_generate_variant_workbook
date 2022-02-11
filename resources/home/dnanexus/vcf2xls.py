import argparse
from pathlib import Path
import re
from string import ascii_uppercase as uppercase
import sys
from typing import Union

import Levenshtein as levenshtein
import numpy as np
from openpyxl.styles import Alignment, Border, colors, DEFAULT_FONT, Font, Side
from openpyxl.styles.fills import PatternFill
import pandas as pd

# openpyxl style settings
THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

class vcf():
    """
    Functions to handle reading and manipulating vcf data

    Called in the order:

    read() -> filter() -> drop_columns() -> reorder() -> merge() -> rename()

    Attributes
    ----------
    args : argparse.Namespace
        arguments passed from command line
    refs : list
        list of genome reference files used for given VCFs
    total_vcf_rows : int
        value to keep total number of rows read in to ensure we don't drop
        any unless --filter is used and --keep is not and => intentionally
        dropping filtered rows
    expanded_vcf_rows : int
        value to track total rows expanded out when multiple transcript
        annotations for one variant are present, resulting in one row per
        transcript annotation per variant in resultant dataframe
    dtypes : dict
        common columns present in annotated vcfs and appropriate dtype to apply
    vcfs : list of pd.DataFrame
        list of dataframes read in from self.args.vcfs
    filtered_rows : pd.DataFrame
        dataframe of all rows dropped from all vcfs
    """

    def __init__(self, args) -> None:
        self.args = args
        self.refs = []
        self.total_vcf_rows = 0
        self.expanded_vcf_rows = 0
        self.filtered_rows = pd.DataFrame()
        self.dtypes = {
            "CHROM": str,
            "POS": int,
            "ID": str,
            "REF": str,
            "ALT": str,
            "QUAL": str,
            "FILTER": str,
            "FORMAT": str,
            "DP": int,
            "SYMBOL": str,
            "VARIANT_CLASS": str,
            "Consequence": str,
            "EXON": str,
            "HGVSc": str,
            "HGVSp": str,
            "gnomAD_AF": float,
            "gnomADg_AF": float,
            "CADD_PHRED": float,
            "Existing_variation": str,
            "ClinVar": str,
            "ClinVar_CLNDN": str,
            "ClinVar_CLNSIG": str,
            "COSMIC": str,
            "Feature": str,
            "STR": bool,
            "RU": bool,
            "Prev_AC": pd.Int16Dtype(),
            "Prev_NS": pd.Int16Dtype()
        }

        # read in the vcfs
        self.vcfs = [self.read(x) for x in args.vcfs]

        print(f"\nTotal variants from {len(self.vcfs)} vcf(s): {self.total_vcf_rows}\n")
        if self.expanded_vcf_rows > 0:
            print(f"Total rows expanded from vcfs: {self.expanded_vcf_rows}")

        if args.print_columns:
            self.print_columns()

        if args.filter:
            # filters vcfs against passed parameters
            self.validate_filters()
            self.build_filters()
            self.filter()

        if args.exclude or args.include:
            self.drop_columns()

        if args.reorder:
            self.order_columns()

        if args.merge:
            self.merge()

        self.rename_columns()

        self.remove_nan()

        # run checks to ensure we haven't unintentionally dropped variants
        self.verify_totals()

        print("\nSUCCESS: Finished munging variants from vcf(s)\n")


    def read(self, vcf) -> pd.DataFrame:
        """
        Reads given vcf into pd.DataFrame, calls following methods:

        - self.parse_header()
        - self.parse_csq_fields()
        - self.split_info()
        - self.split_csq()

        Parameters
        ------
        vcf : str
            path to vcf file to use

        Returns
        -------
        vcf_df : pandas.DataFrame
            dataframe of all variants
        header : list
            vcf header lines
        """
        sample = Path(vcf).stem.split('_')[0]
        print(f"\nReading in vcf {vcf}\n")

        header, columns = self.parse_header(vcf)
        self.get_reference(header)
        csq_fields = self.parse_csq_fields(header)

        # read vcf into pandas df
        vcf_df = pd.read_csv(
            vcf, sep='\t', comment='#', names=columns,
            dtype=self.dtypes, compression='infer'
        )

        self.total_vcf_rows += len(vcf_df.index)  # update our total count
        print(f"Total rows in current VCF: {len(vcf_df.index)}")
        print(f"Total rows of all vcfs read in: {self.total_vcf_rows}\n")

        if self.args.add_name:
            # add sample name from filename as 1st column
            vcf_df.insert(loc=0, column='sampleName', value=sample)

        # split out INFO column values and CSQ
        vcf_df = self.split_info(vcf_df)
        vcf_df = self.split_csq(vcf_df, csq_fields)
        vcf_df = self.split_format_fields(vcf_df)

        # drop INFO and CSQ as we fully split them out
        vcf_df.drop(['INFO', 'CSQ'], axis=1, inplace=True)

        vcf_df = self.set_types(vcf_df)

        return vcf_df


    def parse_header(self, vcf) -> Union[list, list]:
        """
        Read in header lines of given vcf to list, returning the list and the
        vcf column names

        Parameters
        ----------
        vcf : str
            vcf filename to read header from

        Returns
        -------
        header : list
            list of header lines read from vcf
        columns : list
            column names from vcf
        """
        with open(vcf) as fh:
            # read in header of vcf
            header = []
            for line in fh.readlines():
                if line.startswith('#'):
                    header.append(line.rstrip('\n'))
                else:
                    break

        columns = [x.strip('#') for x in header[-1].split()]
        columns[-1] = 'SAMPLE'

        return header, columns


    def get_reference(self, header) -> None:
        """
        Parse reference file used from VCF header

        Parameters
        ----------
        header : list
            lines of vcf header
        """
        ref = next(
            iter([x for x in header if x.startswith('##reference')]), None
        )
        if ref:
            if not ref in self.refs:
                # add reference file if found and same not already in list
                self.refs.append(Path(ref).name)


    def parse_csq_fields(self, header) -> list:
        """
        Parse out csq field names from vcf header.

        Kind of horrible way to parse the CSQ field names but this is how VEP
        seems to have always stored them (6+ years +) in the header as:
        ['##INFO=<ID=CSQ,Number=.,Type=String,Description="Consequence \
        annotations from Ensembl VEP. Format: SYMBOL|VARIANT_CLASS|...

        Parameters
        ----------
        header : list
            list of header lines read from vcf

        Returns
        -------
        csq_fields : list
            list of CSQ field names parsed from vcf header, used to assign as
            column headings when splitting out CSQ data for each variant
        """
        csq_fields = [x for x in header if x.startswith("##INFO=<ID=CSQ")]
        csq_fields = csq_fields[0].split("Format: ")[-1].strip('">').split('|')

        return csq_fields


    def split_format_fields(self, vcf_df) -> pd.DataFrame:
        """
        Get format fields from FORMAT column to split out sample values to
        individual columns

        Parameters
        ----------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf

        Returns
        -------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf with split out FORMAT fields
        """
        # get unique list of FORMAT fields from all rows
        fields = list(set(':'.join(vcf_df.FORMAT.tolist()).split(':')))

        # split out FORMAT and SAMPLE columns a list of ':' joined pairs
        tmp_df = pd.DataFrame()
        tmp_df['tmp'] = vcf_df.apply(
            lambda x: {
                '='.join(x) for x in zip(x.FORMAT.split(':'), x.SAMPLE.split(':'))
            }, axis=1
        )
        tmp_df = tmp_df.astype(str)

        # split every row to a dict of key value pairs and build df
        format_cols = tmp_df.join(pd.DataFrame([
            dict(l) for l in tmp_df.pop('tmp').str.findall((r'(\w+)=([^,\}]+)'))
        ]))

        # some columns have an extra quote added for some reason...
        for col in format_cols.columns:
            format_cols[col] = format_cols[col].apply(
                lambda x: x.rstrip("'") if type(x) == str else x
            )

        # add split out columns to main df
        vcf_df = vcf_df.join(format_cols, rsuffix=" (FMT)")
        vcf_df.drop(['FORMAT', 'SAMPLE'], axis=1, inplace=True)

        return vcf_df


    def split_info(self, vcf_df) -> pd.DataFrame:
        """
        Splits out the INFO column of vcf to all separate values

        Parameters
        ----------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf

        Returns
        -------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf with separated INFO fields
        """
        # get the key value pairs of INFO data
        info_pairs = [
            x.split('CSQ=')[0].split(';') for x in vcf_df['INFO'].tolist()
        ]
        info_pairs = [[x for x in lst if x] for lst in info_pairs]

        # get unique list of keys from key=value pairs in INFO
        info_keys = sorted(list(set([
            x.split('=')[0] if '=' in x else x for pair in info_pairs for x in pair
        ])))
        info_keys = [x for x in info_keys if x]  # can end up with empty string

        info_values = []

        for variant in info_pairs:
            # for every variants values, split them out to dict to add to df
            pair_values = {}

            for value in variant:
                if '=' in value:
                    # key value pair
                    key, value = value.split('=')
                else:
                    # Flag value present (e.g RU, STR)
                    key, value = key, True

                pair_values[key] = value

            info_values.append(pair_values)

        # build df of values to add to main df
        info_df = pd.DataFrame(
            info_values, columns=info_keys
        )

        for col in info_keys:
            # add all info values to main vcf df
            vcf_df[col] = info_df[col]

        return vcf_df


    def split_csq(self, vcf_df, csq_fields) -> pd.DataFrame:
        """
        Split out CSQ string to separate fields to get annotation

        Parameters
        ----------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf
        csq_fields : list
            list of CSQ field names parsed from vcf header, used to assign as
            column headings when splitting out CSQ data for each variant

        Returns
        -------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf with separated CSQ fields
        """
        df_rows = len(vcf_df.index)
        vcf_df['CSQ'] = vcf_df['INFO'].apply(lambda x: x.split('CSQ=')[-1])

        # variants with multiple transcript annotation will have duplicate CSQ
        # data that is comma sepparated => expand this to multiple rows, if no
        # ',' present rows will remain unaffacted (i.e. one transcript)
        columns = list(vcf_df.columns)
        columns.remove('CSQ')

        # set index to be everything except CSQ, expand this to multiple rows
        # then set index back
        vcf_df = vcf_df.set_index(columns).apply(
            lambda x: x.str.split(',').explode()
        ).reset_index()

        # split each CSQ value to own columns
        vcf_df[csq_fields] = vcf_df.CSQ.str.split('|', expand=True)

        if 'COSMIC' in vcf_df.columns:
            # handle known bug in VEP annotation where it duplicates COSMIC
            vcf_df['COSMIC'] = vcf_df['COSMIC'].apply(
                lambda x: '&'.join(set(x.split('&')))
            )

        if df_rows != len(vcf_df.index):
            # total rows has changed => we must have multiple transcripts
            print(f"Total rows of VCF changed on splitting CSQ values")
            print(f"Total rows before: {df_rows}")
            print(f"Total rows after: {len(vcf_df.index)}")
            self.expanded_vcf_rows += len(vcf_df.index) - df_rows

        return vcf_df


    def set_types(self, vcf_df) -> pd.DataFrame:
        """
        Sets appropriate dtypes on given df of variants

        Parameters
        ----------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf

        Returns
        -------
        vcf_df : pd.DataFrame
            dataframe of all variants from a vcf with dtypes set
        """
        # first set any empty strings to pd.NA values to not break types
        vcf_df = vcf_df.replace('', np.nan)

        # get any AF and AC columns that should be floats
        int_columns = [
            x for x in vcf_df.columns if '_AF' in x or '_AC' in x
        ]
        for col in int_columns:
            self.dtypes[col] = float

        # filter all dtypes to just those columns in current df
        df_dtypes = {
            k: v for k, v in self.dtypes.items() if k in list(vcf_df.columns)
        }

        return vcf_df.astype(df_dtypes, errors='ignore')


    def remove_nan(self) -> None:
        """
        Remove NaN values from all dataframes, cast everything to a string
        first as this method is called before writing and it reliably allows
        us to identify and replace all NaN values
        """
        for idx, vcf in enumerate(self.vcfs):
            vcf = vcf.astype(str)

            for col in vcf.columns:
                vcf[col] = vcf[col].apply(
                    lambda x: x.replace('nan', '') if x == 'nan' else x
                )

            self.vcfs[idx] = vcf


    def validate_filters(self) -> None:
        """
        Validate filters passed for filtering variants down.

        These come from args.filter and wil be in the format
        <column><operator><value> (i.e. "Consequence!=synonymous"), currently
        supports the operators >, <, >=, <=, == and !=

        Method will check that a valid operator has been passed, and that the
        given column is present in all columns of all vcfs
        """
        for filter in self.args.filter:
            # check a valid operand passed
            assert len(re.findall('>|<|>=|<=|==|!=', filter)) == 1, (
                f"invalid operand passed in filter: {filter}"
            )

        for vcf in self.vcfs:
            # for each filter, check the specified column is in all the vcfs
            for filter in self.args.filter:
                assert re.split(r'>|<|>=|<=|==|!=', filter)[0] in vcf.columns, (
                    f"Column specified in filter '{filter}' not in vcf "
                    f"columns: {vcf.columns}"
                )


    def build_filters(self) -> None:
        """
        Formats cmd line passed filters as list of lists of each filter
        expression for filtering df of variants.

        This will separate out the passed filters from the format
        "Consequence!=synonymous" -> ["Consequence", "!=", "synonymous"]
        """
        field_value = [
            re.split(r'>|<|>=|<=|==|!=', x) for x in self.args.filter
        ]
        operator = [
            re.findall('>|<|>=|<=|==|!=', x) for x in self.args.filter
        ]

        self.filters = [
            [x[0], y[0], x[1]] for x, y in zip(field_value, operator)
        ]


    def filter(self) -> None:
        """
        Apply filters passed to ech dataframe of variants.

        Filters first checked in self.validate_filters then formatted in
        self.build filters ready to be passed to np.where()

        Currently wrapped in an eval() call which is not ideal but works for
        interpreting the operator passed as a string from the cmd line
        """
        # build list of indices of variants to filter out against specified
        # filters, then apply filter to df, retain filtered rows if --keep set
        if self.args.keep:
            self.filtered_rows = pd.DataFrame()

        for idx, vcf in enumerate(self.vcfs):
            all_filter_idxs = []
            for filter in self.filters:
                col, operator, value = filter[0], filter[1], filter[2]
                if pd.api.types.is_numeric_dtype(vcf[f'{filter[0]}']):
                    # check column we're filtering is numeric and set types
                    value = float(value)
                else:
                    # string values have to be wrapped in quotes from np.where()
                    value = f"'{value}'"

                # get row indices to filter
                filter_idxs = eval((
                    f"np.where(vcf['{col}'].apply("
                    f"lambda x: x {operator} {value}))[0]"
                ))
                all_filter_idxs.extend(filter_idxs)

            # get unique list of indexes matching filters
            all_filter_idxs = sorted(list(set(all_filter_idxs)))

            # apply the filter, assign back the filtered df
            self.filtered_rows = self.filtered_rows.append(
                    vcf.loc[all_filter_idxs], ignore_index=True
                )
            self.vcfs[idx] = vcf.drop(all_filter_idxs)

            filter_string = ', '.join([' '.join(x) for x in self.filters])

            print((
                f"\nApplied the following filters: {filter_string} to vcf(s)\n"
                f"Filtered out {len(all_filter_idxs)} rows\n"
                f"Total rows remaining: {len(self.vcfs[idx].index)}\n\n"
            ))

        self.filtered_rows = self.filtered_rows.reset_index()

        if self.args.keep:
            self.vcfs.append(self.filtered_rows)
            self.args.sheets.append("filtered")


    def print_columns(self) -> None:
        """
        Simple method to just print the columns from each vcf and exit.

        Useful for identify what columns are present in INFO and CSQ fields
        for using --include, --exclude and --reorder arguments
        """
        for name, vcf in zip(self.args.vcfs, self.vcfs):
            print(f"Columns for {Path(name).name}: ")
            print(f"\n\t{list(vcf.columns)}\n\n")

        sys.exit()


    def drop_columns(self) -> None:
        """
        If `--exclude` or `--include` passed, drop given columns
        (or inverse of) from vcf data if they exist.

        If `--include` passed will take the given list of columns and drop the
        remaining columns not specified from all dataframes

        If `--exclude` passed will take the given list of columns and drop
        from all dataframes

        Raises
        ------
        AssertionError
            Raised when columns specified with --include / --exclude are not
            present in one or more of the dataframes
        """
        for idx, vcf in enumerate(self.vcfs):
            if self.args.include:
                # include passed => select all columns not specified to drop
                to_drop = list(
                    set(vcf.columns.tolist()) - set(self.args.include)
                )
            else:
                to_drop = self.args.exclude

            # sense check given exclude columns is in the vcfs
            assert [x for x in vcf.columns for x in to_drop], (
                "Column '{x}' specified with --exclude/--include not "
                "present in one or more of the given vcfs. Valid column "
                f"names: {vcf.columns}"
            )
            self.vcfs[idx].drop(to_drop, axis=1, inplace=True, errors='ignore')


    def order_columns(self) -> None:
        """
        Reorder columns by specified order from `--reorder` argument, any not
        specified will retain original order after reorder columns

        Raises
        ------
        AssertionError
            Raised when columns specified with --reorder are not
            present in one or more of the dataframes
        """
        for idx, vcf in enumerate(self.vcfs):
            vcf_columns = list(vcf.columns)

            # sense check given exclude columns is in the vcfs
            assert [x for x in vcf.columns for x in self.args.reorder], (
                "Column '{x}' specified with --reorder not "
                "present in one or more of the given vcfs. Valid column "
                f"names: {vcf.columns}"
            )

            [vcf_columns.remove(x) for x in self.args.reorder]
            column_order = self.args.reorder + vcf_columns

            self.vcfs[idx] = vcf[column_order]


    def rename_columns(self) -> None:
        """
        Rename columnns from key value pairs passed from --rename argument,
        also remove underscores from all names for nicer reading

        Raises
        ------
        AssertionError
            Raised when columns specified with --rename do not exist in more
            or more of the vcfs columns
        """
        for idx, vcf in enumerate(self.vcfs):
            if self.args.rename:
                # sense check given reorder keys are in the vcfs
                assert [x for x in vcf.columns for x in self.args.rename.keys()], (
                    f"Column(s) specified with --rename not present in one or "
                    f"more of the given vcfs. Valid column names: {vcf.columns}."
                    f"Column names passed to --rename: {self.args.rename.keys()}"
                )
                self.vcfs[idx].rename(columns=dict(self.args.rename.items()))

            # remove underscores from all names
            self.vcfs[idx].columns = [
                x.replace('_', ' ') for x in self.vcfs[idx].columns
            ]


    def merge(self) -> None:
        """
        Merge all variants into one big dataframe, should be used with
        --add_name argument if provenance of variants in merged dataframe
        is important
        """
        self.vcfs = [pd.concat(self.vcfs).reset_index(drop=True)]


    def verify_totals(self) -> None:
        """
        Verify total variants in resultant dataframe(s) match what was read in,
        unless --filter has been applied and --keep has not.

        Totals we have tracked to check on:

        - self.total_vcf_rows -> total rows read in from vcf before modifying
        - self.expanded_rows -> total rows expanded out when CSQ contains
            multiple transcripts, and each transcript for each variant is split
            to individual rows
        - self.filtered_rows -> total rows filtered out to separate df by
            filters passed with --filter
        - total_rows_to_write -> total of all rows that will be written to the
            xlsx across all sheets


        Raises
        ------
        AssertionError
            Raised when the total rows we have tracked don't seem to equal what
            is going to be written to file
        """
        print("Verifying total variants being written to file is correct")
        if not self.args.merge:
            # haven't merged to one df => count all
            total_rows_to_write = sum([len(df.index) for df in self.vcfs])
        else:
            total_rows_to_write = len(self.vcf[0])

        if self.args.filter:
            total_filtered_rows = len(self.filtered_rows)

        # totals we tracked in previous methods
        tracked_total = int(self.total_vcf_rows) + int(self.expanded_vcf_rows)

        print(f"\nTotal variants identified:")
        print(f"\tTotal rows read in from vcf(s): {self.total_vcf_rows}")
        print((
            f"\tTotal rows expanded: "
            f"{self.expanded_vcf_rows}"
        ))
        print(f"\tTotal rows filtered with --filter: {len(self.filtered_rows)}")
        if not self.args.keep:
            print("\t\t--keep not passed, these filtered rows will be dropped")
        else:
            print("\t\t--keep passed, filtered variants will be written to file")
        print(f"\tTotal rows going to write to file: {total_rows_to_write}")
        print(f"\tTotal rows we have tracked: {tracked_total}")


        if not self.args.keep:
            # not keeping filtered rows => check that total we would write if
            # they were included is what we expect from reading in and expanding
            total_rows_to_write += int(len(self.filtered_rows))

        assert tracked_total == total_rows_to_write, (
            "Total rows to be written to file don't appear to equal what has "
            "been tracked, this suggests we have dropped some variants..."
        )


class excel():
    """
    Functions for wrangling variant data into spreadsheet formatting and
    writing output file

    Attributes
    ----------
    args : argparse.Namespace
        arguments passed from command line
    vcfs : list of pd.DataFrame
        list of dataframes formatted to write to file from vcf() methods
    refs : list
        list of reference names parsed from vcf headers
    writer : pandas.io.excel._openpyxl.OpenpyxlWriter
        writer object for writing Excel data to file
    workbook : openpyxl.workbook.workbook.Workbook
        openpyxl workbook object for interacting with per-sheet writing and
        formatting of output Excel file

    Outputs
    -------
    {args.output}.xlsx : file
        Excel file with variants written to, name passed from command line or
        inferred from input vcf name if not specified
    """
    def __init__(self, args, vcfs, refs) -> None:
        print(f"Writing to output file: {Path(args.output).absolute()}")
        self.args = args
        self.vcfs = vcfs
        self.refs = refs
        self.writer = pd.ExcelWriter(args.output, engine='openpyxl')
        self.workbook = self.writer.book

        self.write_summary()
        self.write_variants()
        self.set_font()

        self.workbook.save(args.output)


    def write_summary(self) -> None:
        """
        Write summary sheet to excel file
        """
        if self.args.summary == 'dias':
            # generate summary sheet in format for RD/dias
            self.summary = self.workbook.create_sheet('summary')
            self.dias_summary()


    def dias_summary(self) -> None:
        """
        Write summary sheet in format for RD group, adds the following info:
            - sample ID, panel(s), run IDs etc.
            - formatted tables for them to fill in reporting
        """
        # write titles for summary values
        self.summary.cell(1, 1).value = "Sample ID:"
        self.summary.cell(1, 5).value = "Clinical Indication(s):"
        self.summary.cell(2, 5).value = "Panel(s):"
        self.summary.cell(40, 1).value = "Reads:"
        self.summary.cell(41, 1).value = "Usable Reads:"
        self.summary.cell(43, 1).value = "Workflow:"
        self.summary.cell(44, 1).value = "Workflow ID:"

        # write summary values
        self.summary.cell(1, 2).value = self.args.sample
        self.summary.cell(1, 6).value = self.args.clinical_indication
        self.summary.cell(2, 6).value = self.args.panel
        self.summary.cell(40, 2).value = self.args.reads
        self.summary.cell(41, 2).value = self.args.usable_reads
        self.summary.cell(43, 2).value = self.args.workflow[0]
        self.summary.cell(44, 2).value = self.args.workflow[1]

        # write center reporting section tables
        self.summary.cell(9, 2).value = "Phenotype:"

        self.summary.cell(16, 2).value = "Panels"
        self.summary.cell(16, 3).value = "Excel file"
        self.summary.cell(16, 4).value = "Comments"
        self.summary.cell(16, 6).value = "Analysis by"
        self.summary.cell(16, 7).value = "Date"
        self.summary.cell(16, 8).value = "Checked by"
        self.summary.cell(16, 9).value = "Date"

        self.summary.cell(21, 2).value = "Sanger sequencing confirmation"
        self.summary.cell(22, 2).value = "Gene"
        self.summary.cell(22, 3).value = "NM_#"
        self.summary.cell(22, 4).value = "Coordinate"
        self.summary.cell(22, 5).value = "cDNA"
        self.summary.cell(22, 6).value = "Protein change"
        self.summary.cell(22, 7).value = "WS#"
        self.summary.cell(22, 8).value = "Confirmed (Y/N)"

        self.summary.cell(28, 2).value = "GEM comments summary"
        self.summary.cell(28, 4).value = "Date"

        # merge some title columns that have longer text
        self.summary.merge_cells(start_row=9, end_row=9, start_column=2, end_column=5)
        self.summary.merge_cells(start_row=21, end_row=21, start_column=2, end_column=8)
        self.summary.merge_cells(start_row=16, end_row=16, start_column=4, end_column=5)
        self.summary.merge_cells(start_row=28, end_row=28, start_column=2, end_column=3)
        self.summary.merge_cells(start_row=28, end_row=28, start_column=4, end_column=6)

        # set titles to bold
        title_cells = [
            "A1", "A40", "A41", "A43", "A44", "B1", "B9", "B16", "B21", "B22",
            "B28", "B40", "B41", "B43", "B44", "C16", "C22", "D16", "D22",
            "D28", "E1", "E2", "E22", "F1", "F2", "F16", "F22", "G16", "G22",
            "H16", "H22", "I16"
        ]
        for cell in title_cells:
            self.summary[cell].font = Font(bold=True)

        # set column widths for readability
        self.summary.column_dimensions['A'].width = 13
        self.summary.column_dimensions['B'].width = 13
        self.summary.column_dimensions['C'].width = 13
        self.summary.column_dimensions['D'].width = 13
        self.summary.column_dimensions['E'].width = 18
        self.summary.column_dimensions['F'].width = 16
        self.summary.column_dimensions['G'].width = 16
        self.summary.column_dimensions['H'].width = 16

        # colour title cells
        blueFill = PatternFill(
            patternType="solid", start_color="0CABA8")

        colour_cells =[
            "B9", "B16", "B21", "B22", "B28", "C16", "C22", "D16", "D22",
            "D28", "E22", "F16", "F22", "G16", "G22", "H16", "H22", "I16"
        ]
        for cell in colour_cells:
            self.summary[cell].fill = blueFill

        # set borders around table areas
        row_ranges = [
            'B9:E9', 'B10:E10', 'B11:E11', 'B12:E12', 'B13:E13',
            'B16:I16', 'B17:I17', 'B18:I18',
            'B21:H21', 'B22:H22', 'B23:H23', 'B24:H24', 'B25:H25',
            'B28:F28', 'B29:F29', 'B30:F30', 'B31:F31', 'B32:F32'
        ]
        for row in row_ranges:
            for cells in  self.summary[row]:
                for cell in cells:
                    cell.border = THIN_BORDER


    def write_variants(self) -> None:
        """
        Writes all variants from dataframe(s) to sheet(s) specified in
        self.args.sheets.

        If sheet names not specified, these will be set as "variant" where one
        dataframe is being written, or the vcf filename prefix if there are
        more than one dataframes to write
        """
        total_rows = sum([len(x) for x in self.vcfs])
        print(f"Writing {total_rows} rows to output xlsx file")
        with self.writer:
            # add variants
            for sheet, vcf in zip(self.args.sheets, self.vcfs):
                vcf.to_excel(
                    self.writer, sheet_name=sheet, index=False
                )
                curr_worksheet = self.writer.sheets[sheet]
                self.set_widths(curr_worksheet, vcf.columns)
                self.workbook.save(self.args.output)


    def set_font(self) -> None:
        """
        Set font to all cells in sheet to Calibri

        Default is Times New Roman and it is ugly
        """
        for ws in self.workbook:
            for cells in ws.rows:
                for cell in cells:
                    cell.font = Font(name="Calibri")


    def set_widths(self, current_sheet, sheet_columns) -> None:
        """
        Set widths for variant sheets off common names to be more readable

        Parameters
        ----------
        current_sheet : openpyxl.Writer
            writer object for current sheet
        sheet_columns : list
            column names for sheet from DataFrame.columns
        """
        widths = {
            "chrom": 8,
            "pos": 12,
            "ref": 10,
            "alt": 10,
            "qual": 10,
            "af": 6,
            "dp": 10,
            'ac': 10,
            'af': 10,
            'an': 10,
            'dp': 10,
            'baseqranksum': 15,
            'clippingranksum': 16,
            "symbol": 12,
            "exon": 9,
            "variant class": 15,
            "consequence": 25,
            "hgvsc": 24,
            "hgvsp": 24,
            "gnomad": 13,
            "existing variation": 18,
            "clinvar": 10,
            "clinvar clndn": 18,
            "clinvar clinsig": 18,
            "cosmic": 15,
            "feature": 17
        }

        # generate list of 78 potential xlsx columns from A,B,C...BX,BY,BZ
        # allows for handling a lot of columns
        column_list = [
            f"{x}{y}" for x in ['', 'A', 'B'] for y in uppercase
        ]

        for idx, column in enumerate(sheet_columns):
            # loop over column names, select width by closest match from dict
            # and set width in sheet letter
            width = self.get_closest_match(column.lower(), widths)
            current_sheet.column_dimensions[column_list[idx]].width = width


    def get_closest_match(self, column, widths) -> int:
        """
        Given a column name, find the closest match (if there is one) in the
        widths dict and return its width value to set. Using imprecise name
        matching as columns can differ between variant callers, annotation in
        VEP and renaming in vcf.rename_columns()

        Parameters
        ----------
        column : str
            name of column
        widths : dict
            dict of common column names and widths

        Returns
        -------
        width : int
            column width value to set
        """
        distances = {x: levenshtein.distance(column, x) for x in widths.keys()}
        closest_match = min(distances, key=distances.get)

        if distances[closest_match] <= 5:
            # close enough match to probably be correct
            width = widths[closest_match]
        else:
            # no close matches to name, use default width
            width = 13

        return width


class parsePairs(argparse.Action):
    """
    Simple class method for enabling passing of key value pairs to argparse
    as this is not natively supported by argparse

    Used for passing in names to rename columns in output Excel
    """
    def __call__(self, parser, namespace, values, option_string=None):
        setattr(namespace, self.dest, dict())
        for value in values:
            key, value = value.split('=')
            getattr(namespace, self.dest)[key] = value


def parse_args() -> argparse.Namespace:
    """
    Parse command line arguments

    Returns
    -------
    args : Namespace
        Namespace of passed command line argument inputs
    """
    parser = argparse.ArgumentParser(
        'Turns an annotated vcf into an xlsx for human viewing'
    )

    parser.add_argument(
        '-v', '--vcfs', nargs='+',
        help='Annotated vcf file'
    )
    parser.add_argument(
        '-e', '--exclude', nargs='+',
        help=(
            'Columns in vcf to EXCLUDE from output, by default all INFO and '
            'CSQ fields are expanded to their own columns'
        )
    )
    parser.add_argument(
        '-i', '--include', nargs='+',
        help=(
            'Columns in vcf to INCLUDE from output, by default all INFO and '
            'CSQ fields are expanded to their own columns'
        )
    )
    parser.add_argument(
        '-r', '--reorder', required=False, nargs='+',
        help=(
            'Set order for columns in output vcf, any not specified will be '
            'appended to the end'
        )
    )
    parser.add_argument(
        '-z', '--rename', nargs='*', action=parsePairs,
        help=(
            'Pass pairs of {column_name}={new_column_name} for renaming '
            'columns in output excel, should be passed as '
            '--rename CHROM=chr POS=pos REF=ref...'
        )
    )
    parser.add_argument(
        '-f', '--filter', nargs='+',
        help=(
            'Columns on which to filter out variants. Format should be '
            'as <column><operator><value> (gnomAD_AF<0.02). Supported '
            'operands are >|<|>=|<=|==|!='
        )
    )
    parser.add_argument(
        '-k', '--keep', action='store_true',
        help=(
            'Pass when using --filter to keep filtered variants in a '
            'separated "filtered" tab'
        )
    )
    parser.add_argument(
        '-n', '--add_name', action='store_true',
        help='Add sample name from filename as first column'
    )
    parser.add_argument(
        '-s', '--sheets', nargs='+',
        help=(
            'Names to use for multiple sheets, these MUST be the same number '
            'as the number of vcfs passed and in the same order. If not '
            'given, if there is 1 vcf passed the sheet will be named '
            '"variants", if multiple the name prefix of the vcf will be used'
        )
    )
    parser.add_argument(
        '-o', '--output', required=False,
        help=(
            'output name prefix for file, if more than 1 vcf passed a name '
            'must be specified. If only 1 vcf passed and no output name, the '
            'vcf filename prefix will be used'
        )
    )
    parser.add_argument(
        '-m', '--merge', action='store_true',
        help='Merge multiple vcfs into one dataframe of variants to display'
    )
    parser.add_argument(
        '--summary', required=False,
        help='summary sheet to include, must be one of: dias'
    )
    parser.add_argument(
        '--analysis', default='',
        help='Name of analysis to display in summary'
    )
    parser.add_argument(
        '--workflow', default=('', ''), nargs=2,
        help='ID and name of workflow to display in summary'
    )
    parser.add_argument(
        '--panel', default='',
        help='panel name to display in summary'
    )
    parser.add_argument(
        '--clinical_indication', default='',
        help="clinical indication to write into summary sheet"
    )
    parser.add_argument(
        '--sample', default='',
        help='name of sample to display in summary report'
    )
    parser.add_argument(
        '--reads', default='',
        help='total reads from flagstat'
    )
    parser.add_argument(
        '--usable_reads', default='',
        help='usable reads from flagstat'
    )
    parser.add_argument(
        '--print-columns', required=False, action='store_true',
        help=(
            'Print total columns of all vcfs that will be output to the xlsx. '
            'Useful to identify what will be in the output to include / exclude'
        )
    )

    args = parser.parse_args()

    if not args.output:
        if len(args.vcfs) > 1:
            raise RuntimeError((
                "More than one vcf passed but no output name specified with "
                "--output"
            ))
        else:
            args.output = Path(args.vcfs[0]).name.replace(
                '.vcf', '').replace('.gz', '')

    args.output += ".xlsx"

    if not args.sheets:
        if len(args.vcfs) > 1 and not args.merge:
            # sheet names not specified for > 1 vcf passed => use vcf names
            args.sheets = [Path(x).name.split('_')[0] for x in args.vcfs]
        else:
            # one vcf => name it variants
            args.sheets = ["variants"]
    else:
        assert len(args.vcfs) == len(args.sheets), (
            "Different number of sheets specified to total vcfs passed. Number of "
            f"vcf passed: {len(args.vcfs)}. Number of sheet names passed: "
            f"{len(args.sheets)}"
        )

    return args


def main():
    args = parse_args()
    vcf_handler = vcf(args)
    excel(args, vcf_handler.vcfs, vcf_handler.refs)


if __name__ == "__main__":
    main()
